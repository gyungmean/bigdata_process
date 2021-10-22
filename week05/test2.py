#!/usr/bin/python3

import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(filename = "student.xlsx")
sheet_ranges = wb["Sheet1"]

row_num = 2
while True:
	total = sheet_ranges.cell(row = row_num, column = 7).value
	if total == None:
		break
	print(total)

	col_total = col_midterm * 0.3 + col_final * 0.35 + col_homework * 0.34 + col_attendance
	print(col_total)
#sheet_ranges.cell(row = row_num, column = 7, value = col_total)

#	eval_string = "=0.3*{}{}+0.35*{}{}+0.34*{}{}+{}{}".format(get_column_letter(3), row_num, get_column_letter(4), row_num, get_column_letter(5), row_num, get_column_letter(6), row_num)
#	sheet_ranges.cell(row = row_num, column = 7, value = eval_string)
 	
	row_num = row_num + 1


wb.save(filename = "student.xlsx")
