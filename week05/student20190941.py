#!/usr/bin/python3

import openpyxl
from openpyxl.utils import get_column_letter
import math

wb = openpyxl.load_workbook(filename = "student.xlsx")
sheet_ranges = wb["Sheet1"]

row_num = 2
while True:
	col_midterm = sheet_ranges.cell(row = row_num, column = 3).value
	col_final = sheet_ranges.cell(row = row_num, column = 4).value
	col_homework = sheet_ranges.cell(row = row_num, column = 5).value
	col_attendance = sheet_ranges.cell(row = row_num, column = 6).value
	if col_midterm == None:
		break
	
	col_total = col_midterm * 0.3 + col_final * 0.35 + col_homework * 0.34 + col_attendance
	sheet_ranges.cell(row = row_num, column = 7, value = col_total)

	#eval_string = "=0.3*{}{}+0.35*{}{}+0.34*{}{}+{}{}".format(get_column_letter(3), row_num, get_column_letter(4), row_num, get_column_letter(5), row_num, get_column_letter(6), row_num)
	#sheet_ranges.cell(row = row_num, column = 7, value = eval_string)
 	
	row_num = row_num + 1

list = [] #for sort
row_num = 2
while True:
	col_total = sheet_ranges.cell(row = row_num, column = 7).value
	
	if col_total == None:
		break

	list.append((row_num, col_total))	

	row_num = row_num + 1 	

list.sort(key=lambda x:x[1], reverse=True)

row_num = row_num - 2

a_rank = math.floor(row_num * 0.3)
b_rank = math.floor(row_num * 0.7)

index = 0
while index < row_num:
	if int(list[index][1]) < 40:
		sheet_ranges.cell(row = list[index][0], column = 8, value="F")
	index = index + 1

index = 0
while True:
	plus = math.floor(a_rank * 0.5)
	while index < plus:
		sheet_ranges.cell(row = list[index][0], column = 8, value="A+")
		index = index + 1

	while index < a_rank:
		sheet_ranges.cell(row = list[index][0], column = 8, value="A0")
		index = index + 1

	break

while True:
	plus = (b_rank - a_rank) / 2 + index
	while index < plus:
		sheet_ranges.cell(row = list[index][0], column = 8, value="B+")
		index = index + 1

	while index < b_rank:
		sheet_ranges.cell(row = list[index][0], column = 8, value="B0")
		index = index + 1	

	break

while True:
	plus = (row_num - b_rank) / 2 + index
	while index < plus:
		value = sheet_ranges.cell(row = list[index][0], column = 8).value
		if value == None:
			sheet_ranges.cell(row = list[index][0], column = 8, value="C+")
		index = index + 1
	while index < row_num:
		value = sheet_ranges.cell(row = list[index][0], column = 8).value
		if value == None:
			sheet_ranges.cell(row = list[index][0], column = 8, value="C0")	
		index = index + 1

	break

wb.save(filename = "student.xlsx")
